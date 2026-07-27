#!/usr/bin/env python3
# 将 002-高频词汇库.md 解析为可直接导入「腾讯文档·多维表格」的 Excel 文件
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SRC = r"D:\华的obsidian\210-写作提高\002-知识图谱\002-高频词汇库.md"
OUT = r"C:\Users\Lenovo\WorkBuddy\2026-07-21-16-55-18\vocab-checkin\高频词汇打卡-腾讯多维表格导入.xlsx"

def strip_md(s: str) -> str:
    s = s.strip()
    s = re.sub(r'^\*\*(.*?)\*\*$', r'\1', s)  # 去加粗
    s = re.sub(r'`(.*?)`', r'\1', s)          # 去行内代码
    return s.strip()

with open(SRC, encoding='utf-8') as f:
    lines = f.readlines()

rows = []
cur_cat = ''
for ln in lines:
    if ln.startswith('## '):
        cur_cat = ln[3:].strip()
        continue
    if not ln.lstrip().startswith('|'):
        continue
    cells = [c.strip() for c in ln.strip().strip('|').split('|')]
    if not cells:
        continue
    # 只要首列加粗的词条行（与源文件 686 条口径一致），跳过表头/分隔行
    if not cells[0].startswith('**'):
        continue
    if '高级表达' in cells[0]:
        continue
    if len(cells) < 5:
        cells = cells + [''] * (5 - len(cells))
    word = strip_md(cells[0])
    if not word:
        continue
    rows.append({
        'cat': cur_cat,
        'word': word,
        'syn': strip_md(cells[1]),
        'meaning': strip_md(cells[2]),
        'example': strip_md(cells[3]),
        'scene': strip_md(cells[4]),
    })

print(f"解析到词条数: {len(rows)}")
print(f"分类数: {len(set(r['cat'] for r in rows))}")
for c in sorted(set(r['cat'] for r in rows)):
    print("  -", c)

# ---- 构建工作簿 ----
wb = Workbook()

# 说明 sheet（放最前，导入时可删）
ws_info = wb.active
ws_info.title = "说明（导入后可删）"
info_rows = [
    ["高频词汇打卡 · 腾讯文档多维表格 导入文件"],
    [""],
    ["【字段与导入后建议设置的类型】"],
    ["分类", "导入后改为「单选」，便于分组/筛选"],
    ["高级表达", "文本（主字段，建议设为主列）"],
    ["口语化同义词", "文本"],
    ["释义", "文本"],
    ["例句", "文本"],
    ["使用场景", "文本"],
    ["打卡次数", "数字，默认 0（按钮会自动 +1）"],
    ["首次使用", "日期（空，首次打卡时由按钮写入）"],
    ["最近使用", "日期（空，每次打卡由按钮写入）"],
    ["掌握度", "单选：未用 / 偶尔 / 熟练（按钮可同步升级）"],
    [""],
    ["【打卡按钮配置】"],
    ["字段 → 按钮 → 命名“打卡+1” → 自动化："],
    ["  触发=点击按钮；操作=修改记录"],
    ["  ① 打卡次数 = 打卡次数 + 1"],
    ["  ② 最近使用 = 当前时间"],
    ["  ③ 若 首次使用 为空 → 首次使用 = 当前时间"],
    [""],
    ["【分组/筛选】"],
    ["分组：工具栏「分组」→ 按 分类"],
    ["筛选：工具栏「筛选」→ 打卡次数 = 0（看未打卡词）"],
    ["手机：腾讯文档 App 打开本表，点按钮即可打卡"],
]
for r in info_rows:
    ws_info.append(r)
ws_info['A1'].font = Font(bold=True, size=14)
for i in range(3, len(info_rows) + 1):
    if info_rows[i-1] and not info_rows[i-1][0].startswith(' ') and info_rows[i-1][0] and '：' not in info_rows[i-1][0] and '→' not in info_rows[i-1][0]:
        ws_info.cell(row=i, column=1).font = Font(bold=True)

# 数据 sheet
ws = wb.create_sheet("打卡表")
headers = ["分类", "高级表达", "口语化同义词", "释义", "例句", "使用场景",
           "打卡次数", "首次使用", "最近使用", "掌握度"]
ws.append(headers)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF")
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")

for r in rows:
    ws.append([r['cat'], r['word'], r['syn'], r['meaning'], r['example'],
               r['scene'], 0, "", "", "未用"])

# 列宽
widths = [16, 18, 18, 30, 40, 24, 9, 12, 12, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

wb.save(OUT)
print("已生成:", OUT)
